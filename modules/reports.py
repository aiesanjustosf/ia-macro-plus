import io
from datetime import datetime

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle


BANCO_MACRO_RAZON_SOCIAL = "BANCO MACRO S.A."
BANCO_MACRO_CUIT = "30500010084"
BANCO_MACRO_TD = "80"


def fmt_money(value: float) -> str:
    try:
        return f"{float(value):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except Exception:
        return "0,00"


def _neto(df: pd.DataFrame, mask: pd.Series) -> float:
    # Para gastos/impuestos: débitos negativos suman como costo; créditos positivos restan.
    return -df.loc[mask, "importe"].sum()


def build_masks(df: pd.DataFrame) -> dict[str, pd.Series]:
    concepto = df["concepto_norm"].fillna("")
    m_25413 = concepto.str.contains("25413", na=False)
    m_25413_db = m_25413 & concepto.str.contains("S/DB|S DB|DB TASA", regex=True, na=False)
    m_25413_cr = m_25413 & concepto.str.contains("S/CR|S CR|CR TASA", regex=True, na=False)
    m_sircreb = concepto.str.contains("SIRCREB|IIBB SANTA FE|AJ IIBB|IIBB STAFE", regex=True, na=False)
    m_iva_basico = concepto.str.contains("DEBITO FISCAL IVA BASICO", na=False)
    m_iva_percepcion = concepto.str.contains("DEBITO FISCAL IVA PERCEPCION", na=False)
    m_comisiones = concepto.str.contains("COMISION|MANTENIMIENTO|INTER.ADEL|INTER ADEL", regex=True, na=False) & ~m_iva_basico & ~m_iva_percepcion
    m_afip_arca = concepto.str.contains("AFIP|ARCA", regex=True, na=False)
    m_cheques = concepto.str.contains("CHEQUE", na=False)
    m_tarjetas = concepto.str.contains("TARJETA", na=False)
    m_prestamos = concepto.str.contains("PRESTAMOS|PRESTAMO", regex=True, na=False)
    m_sueldos = concepto.str.contains("PAGO REMUNERACIONES", na=False)
    return {
        "25413": m_25413,
        "25413_db": m_25413_db,
        "25413_cr": m_25413_cr,
        "sircreb": m_sircreb,
        "iva_basico": m_iva_basico,
        "iva_percepcion": m_iva_percepcion,
        "comisiones": m_comisiones,
        "afip_arca": m_afip_arca,
        "cheques": m_cheques,
        "tarjetas": m_tarjetas,
        "prestamos": m_prestamos,
        "sueldos": m_sueldos,
    }


def build_operational_summary(df: pd.DataFrame) -> pd.DataFrame:
    masks = build_masks(df)

    rows = [
        {"Concepto": "Ley 25.413 / Débitos y Créditos - Neto", "Importe": _neto(df, masks["25413"]), "incluye_total": True},
        {"Concepto": "Ley 25.413 S/DB", "Importe": _neto(df, masks["25413_db"]), "incluye_total": False},
        {"Concepto": "Ley 25.413 S/CR", "Importe": _neto(df, masks["25413_cr"]), "incluye_total": False},
        {"Concepto": "SIRCREB / IIBB Santa Fe - Neto", "Importe": _neto(df, masks["sircreb"]), "incluye_total": True},
        {"Concepto": "Gastos / Comisiones bancarias", "Importe": _neto(df, masks["comisiones"]), "incluye_total": True},
        {"Concepto": "IVA básico sobre gastos bancarios", "Importe": _neto(df, masks["iva_basico"]), "incluye_total": True},
        {"Concepto": "IVA percepción", "Importe": _neto(df, masks["iva_percepcion"]), "incluye_total": True},
        {"Concepto": "AFIP / ARCA", "Importe": _neto(df, masks["afip_arca"]), "incluye_total": True},
        {"Concepto": "Cheques", "Importe": _neto(df, masks["cheques"]), "incluye_total": True},
        {"Concepto": "Tarjetas", "Importe": _neto(df, masks["tarjetas"]), "incluye_total": True},
        {"Concepto": "Préstamos", "Importe": _neto(df, masks["prestamos"]), "incluye_total": True},
        {"Concepto": "Sueldos / Remuneraciones", "Importe": _neto(df, masks["sueldos"]), "incluye_total": True},
    ]

    resumen = pd.DataFrame(rows)
    resumen = resumen[resumen["Importe"].round(2) != 0].copy()
    total = resumen.loc[resumen["incluye_total"], "Importe"].sum()
    resumen = resumen.drop(columns=["incluye_total"])
    resumen.loc[len(resumen)] = {"Concepto": "TOTAL RESUMEN OPERATIVO", "Importe": total}
    resumen["Importe formateado"] = resumen["Importe"].apply(fmt_money)
    return resumen


def _style_sheet(ws):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    total_fill = PatternFill("solid", fgColor="D9EAF7")
    thin = Side(style="thin", color="D9D9D9")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = Border(bottom=thin)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = Border(bottom=thin)
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00'
        if row[0].value and "TOTAL" in str(row[0].value).upper():
            for cell in row:
                cell.fill = total_fill
                cell.font = Font(bold=True)

    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            max_len = max(max_len, len(str(cell.value)) if cell.value is not None else 0)
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 55)

    ws.freeze_panes = "A2"


def _fmt_date(value) -> str:
    if pd.isna(value):
        return ""
    if hasattr(value, "strftime"):
        return value.strftime("%d/%m/%Y")
    return str(value)




def _saldo_key(value) -> int:
    """Clave entera en centavos para comparar saldos sin ruido de float."""
    try:
        return int(round(float(value) * 100))
    except Exception:
        return 0


def _ordered_for_reconciliation(group: pd.DataFrame) -> pd.DataFrame:
    """
    Ordena movimientos para conciliación.

    Prioridad:
    1) reconstrucción por cadena de saldos: saldo anterior + importe = saldo actual;
    2) fallback por fecha más antigua a más reciente y, dentro del día, orden inverso al PDF
       porque Macro suele listar de más nuevo a más viejo.

    Esto evita tomar como saldo inicial el primer renglón del PDF cuando el archivo viene
    descendente y ese primer registro corresponde al final del mes.
    """
    g = group.copy().reset_index(drop=False).rename(columns={"index": "_orig_index"})
    if g.empty:
        return g

    g["_saldo_key"] = g["saldo"].apply(_saldo_key)
    g["_prev_key"] = (g["saldo"] - g["importe"]).apply(_saldo_key)

    current_keys = set(g["_saldo_key"].tolist())
    start_candidates = g[~g["_prev_key"].isin(current_keys)].copy()

    if not start_candidates.empty:
        # Si hay más de un candidato, prioriza fecha más antigua.
        sort_cols = ["fecha"] if "fecha" in start_candidates.columns else []
        if "orden_pdf" in start_candidates.columns:
            sort_cols.append("orden_pdf")
            asc = [True, False]
        else:
            asc = [True] * len(sort_cols)
        start = start_candidates.sort_values(sort_cols, ascending=asc).iloc[0] if sort_cols else start_candidates.iloc[0]

        unused = set(g.index.tolist())
        ordered_idx = []
        current_idx = int(start.name)

        while current_idx in unused:
            ordered_idx.append(current_idx)
            unused.remove(current_idx)
            current_saldo = int(g.loc[current_idx, "_saldo_key"])
            candidates = g[(g.index.isin(unused)) & (g["_prev_key"] == current_saldo)].copy()
            if candidates.empty:
                break
            # Si hay empate, toma la fecha siguiente más antigua posible.
            sort_cols = ["fecha"] if "fecha" in candidates.columns else []
            if "orden_pdf" in candidates.columns:
                sort_cols.append("orden_pdf")
                asc = [True, False]
            else:
                asc = [True] * len(sort_cols)
            nxt = candidates.sort_values(sort_cols, ascending=asc).iloc[0] if sort_cols else candidates.iloc[0]
            current_idx = int(nxt.name)

        if len(ordered_idx) == len(g):
            return g.loc[ordered_idx].drop(columns=["_saldo_key", "_prev_key"], errors="ignore").reset_index(drop=True)

    # Fallback explícito pedido: saldo inicial desde fecha más antigua y saldo final desde fecha más reciente.
    sort_cols = []
    ascending = []
    if "fecha" in g.columns:
        sort_cols.append("fecha")
        ascending.append(True)
    if "orden_pdf" in g.columns:
        sort_cols.append("orden_pdf")
        ascending.append(False)
    elif "orden_cronologico" in g.columns:
        sort_cols.append("orden_cronologico")
        ascending.append(True)

    if sort_cols:
        g = g.sort_values(sort_cols, ascending=ascending)

    return g.drop(columns=["_saldo_key", "_prev_key"], errors="ignore").reset_index(drop=True)

def build_bank_reconciliation(df_mov: pd.DataFrame) -> pd.DataFrame:
    """
    Conciliación bancaria para PDFs de Últimos Movimientos.

    Banco Macro no informa saldo anterior en este formato. Se reconstruye así:
    saldo anterior calculado = saldo del primer movimiento cronológico - importe del primer movimiento cronológico.
    Luego se recalcula el saldo acumulado y se compara contra el saldo final informado.
    """
    if df_mov.empty:
        return pd.DataFrame()

    group_cols = []
    if "archivo" in df_mov.columns:
        group_cols.append("archivo")
    if "cuenta" in df_mov.columns:
        group_cols.append("cuenta")

    if not group_cols:
        grouped = [("PDF", df_mov.copy())]
    else:
        grouped = df_mov.groupby(group_cols, dropna=False, sort=False)

    rows = []
    for key, group in grouped:
        g = _ordered_for_reconciliation(group)

        if g.empty:
            continue

        first = g.iloc[0]
        last = g.iloc[-1]
        saldo_anterior = round(float(first["saldo"]) - float(first["importe"]), 2)
        total_creditos = round(g.loc[g["importe"] > 0, "importe"].sum(), 2)
        total_debitos = round(-g.loc[g["importe"] < 0, "importe"].sum(), 2)
        neto = round(g["importe"].sum(), 2)
        saldo_final_calculado = round(saldo_anterior + neto, 2)
        saldo_final_informado = round(float(last["saldo"]), 2)
        diferencia = round(saldo_final_calculado - saldo_final_informado, 2)

        if isinstance(key, tuple):
            key_values = list(key)
        else:
            key_values = [key]

        archivo = ""
        cuenta = "s/n"
        if "archivo" in group_cols and "cuenta" in group_cols:
            archivo, cuenta = key_values[0], key_values[1]
        elif "archivo" in group_cols:
            archivo = key_values[0]
            cuenta = str(g["cuenta"].iloc[0]) if "cuenta" in g.columns else "s/n"
        elif "cuenta" in group_cols:
            cuenta = key_values[0]
            archivo = str(g["archivo"].iloc[0]) if "archivo" in g.columns else ""

        rows.append({
            "Archivo": archivo,
            "Cuenta": cuenta,
            "Desde": _fmt_date(g["fecha"].min()) if "fecha" in g.columns else "",
            "Hasta": _fmt_date(g["fecha"].max()) if "fecha" in g.columns else "",
            "Movimientos": len(g),
            "Saldo anterior calculado": saldo_anterior,
            "Créditos": total_creditos,
            "Débitos": total_debitos,
            "Neto movimientos": neto,
            "Saldo final calculado": saldo_final_calculado,
            "Saldo final informado": saldo_final_informado,
            "Diferencia": diferencia,
            "Estado": "Conciliado" if abs(diferencia) < 0.01 else "Diferencia",
        })

    out = pd.DataFrame(rows)
    for col in [
        "Saldo anterior calculado", "Créditos", "Débitos", "Neto movimientos",
        "Saldo final calculado", "Saldo final informado", "Diferencia"
    ]:
        if col in out.columns:
            out[f"{col} formateado"] = out[col].apply(fmt_money)
    return out


def build_holistor_import(df_mov: pd.DataFrame) -> pd.DataFrame:
    """
    Genera una planilla de importación orientada a Holistor Compras.
    Criterio usado:
    - Un comprobante RB del Banco Macro por cuenta/período.
    - Gastos/comisiones como neto al 21% con IVA básico separado en la misma línea.
    - Ley 25.413 y SIRCREB como conceptos no gravados/percepciones según corresponda.
    """
    masks = build_masks(df_mov)
    fecha = df_mov["fecha"].max()
    fecha_txt = _fmt_date(fecha)
    cuenta = str(df_mov["cuenta"].iloc[0]) if "cuenta" in df_mov.columns and not df_mov.empty else "s/n"
    nro = f"{fecha.strftime('%Y%m%d') if hasattr(fecha, 'strftime') else '00000000'}{cuenta[-4:] if cuenta != 's/n' else '0000'}"

    gasto_comisiones = round(_neto(df_mov, masks["comisiones"]), 2)
    iva_basico = round(_neto(df_mov, masks["iva_basico"]), 2)
    iva_percepcion = round(_neto(df_mov, masks["iva_percepcion"]), 2)
    ley_25413 = round(_neto(df_mov, masks["25413"]), 2)
    sircreb = round(_neto(df_mov, masks["sircreb"]), 2)

    base = {
        "Fecha": fecha_txt,
        "Tipo de comprobante": "RB",
        "Letra": "",
        "Punto de venta": "0001",
        "Número": nro,
        "Razón Social": BANCO_MACRO_RAZON_SOCIAL,
        "TD": BANCO_MACRO_TD,
        "CUIT": BANCO_MACRO_CUIT,
        "Cond. fiscal": "RI",
        "Detalle": "Banco Macro - Últimos Movimientos",
        "Cód. Neto": "",
        "Neto": 0.0,
        "Alícuota IVA": "0,000",
        "IVA": 0.0,
        "Cód": "584",
        "Exento / No Gravado": 0.0,
        "Cód p/R": "",
        "Percepción / Retención": 0.0,
        "Observación": f"Cuenta {cuenta}",
    }

    rows = []

    if gasto_comisiones or iva_basico:
        row = base.copy()
        row.update({
            "Detalle": "Gastos / Comisiones bancarias",
            "Cód. Neto": "506",
            "Neto": gasto_comisiones,
            "Alícuota IVA": "21,000" if iva_basico else "0,000",
            "IVA": iva_basico,
        })
        rows.append(row)

    if ley_25413:
        row = base.copy()
        row.update({
            "Detalle": "Ley 25.413 / Débitos y Créditos",
            "Cód. Neto": "506",
            "Exento / No Gravado": ley_25413,
        })
        rows.append(row)

    if sircreb:
        row = base.copy()
        row.update({
            "Detalle": "SIRCREB / IIBB Santa Fe",
            "Cód p/R": "SIRC",
            "Percepción / Retención": sircreb,
        })
        rows.append(row)

    if iva_percepcion:
        row = base.copy()
        row.update({
            "Detalle": "Percepción IVA",
            "Cód p/R": "P007",
            "Percepción / Retención": iva_percepcion,
        })
        rows.append(row)

    if not rows:
        rows.append(base.copy())

    return pd.DataFrame(rows)



def build_credit_detail(df_mov: pd.DataFrame) -> pd.DataFrame:
    """
    Detalle de créditos/financiaciones detectadas.

    Incluye acreditaciones vinculadas a créditos y débitos de pago de cuotas o financiación.
    No toma todos los ingresos bancarios: solo conceptos con señales de crédito/préstamo/cuota.
    """
    if df_mov.empty:
        return pd.DataFrame()

    c = df_mov["concepto_norm"].fillna("")
    mask = c.str.contains(
        r"CREDIN|PRESTAMO|PRESTAMOS|CUOTA|CIRCULO CERRADO|ADEL\.CC|ACUERD|INTER\.ADEL|PAGO.*CUOTA|DEBITO.*PREST",
        regex=True,
        na=False,
    )

    out = df_mov.loc[mask].copy()
    if out.empty:
        return pd.DataFrame(columns=[
            "archivo", "cuenta", "fecha", "referencia", "causal", "tipo_credito",
            "concepto", "importe", "saldo", "signo", "observacion"
        ])

    def tipo(row):
        concepto = str(row.get("concepto_norm", ""))
        importe = float(row.get("importe", 0) or 0)
        if importe > 0:
            return "Acreditación de crédito"
        if "INTER" in concepto or "ACUERD" in concepto or "ADEL.CC" in concepto:
            return "Intereses / acuerdo"
        return "Pago de cuota / débito"

    out["tipo_credito"] = out.apply(tipo, axis=1)
    out["signo"] = out["importe"].apply(lambda v: "Crédito" if float(v) > 0 else "Débito")
    out["observacion"] = out["tipo_credito"]

    if "fecha" in out.columns:
        out = out.sort_values(["fecha", "orden_pdf"] if "orden_pdf" in out.columns else ["fecha"], ascending=[True, False] if "orden_pdf" in out.columns else [True])

    cols = [
        "archivo", "cuenta", "fecha", "referencia", "causal", "tipo_credito",
        "concepto", "importe", "saldo", "signo", "observacion"
    ]
    cols = [col for col in cols if col in out.columns]
    return out[cols].reset_index(drop=True)


def make_credit_detail_excel(df_mov: pd.DataFrame) -> bytes:
    output = io.BytesIO()
    detalle = build_credit_detail(df_mov)

    export = detalle.copy()
    if not export.empty and "fecha" in export.columns:
        export["fecha"] = export["fecha"].dt.strftime("%d/%m/%Y")

    resumen = pd.DataFrame()
    if not detalle.empty:
        resumen = detalle.groupby("tipo_credito", dropna=False).agg(
            Cantidad=("importe", "count"),
            Total=("importe", "sum"),
        ).reset_index()
        resumen.loc[len(resumen)] = {
            "tipo_credito": "TOTAL",
            "Cantidad": int(resumen["Cantidad"].sum()),
            "Total": float(resumen["Total"].sum()),
        }

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        resumen.to_excel(writer, sheet_name="Resumen creditos", index=False)
        export.to_excel(writer, sheet_name="Detalle creditos", index=False)
        for ws in writer.book.worksheets:
            _style_sheet(ws)

    return output.getvalue()

def make_excel(df_mov: pd.DataFrame, df_resumen: pd.DataFrame, df_conciliacion: pd.DataFrame | None = None) -> bytes:
    output = io.BytesIO()

    df_mov_export = df_mov.copy()
    if "fecha" in df_mov_export.columns:
        df_mov_export["fecha"] = df_mov_export["fecha"].dt.strftime("%d/%m/%Y")

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_resumen.to_excel(writer, sheet_name="Resumen operativo", index=False)
        if df_conciliacion is not None and not df_conciliacion.empty:
            df_conciliacion.to_excel(writer, sheet_name="Conciliacion bancaria", index=False)
        credit_detail = build_credit_detail(df_mov)
        if not credit_detail.empty:
            credit_export = credit_detail.copy()
            if "fecha" in credit_export.columns:
                credit_export["fecha"] = credit_export["fecha"].dt.strftime("%d/%m/%Y")
            credit_export.to_excel(writer, sheet_name="Detalle creditos", index=False)
        df_mov_export.to_excel(writer, sheet_name="Movimientos", index=False)

        for ws in writer.book.worksheets:
            _style_sheet(ws)

    return output.getvalue()


def make_holistor_excel(df_mov: pd.DataFrame) -> bytes:
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        if "cuenta" in df_mov.columns:
            for cuenta, df_cuenta in df_mov.groupby("cuenta", dropna=False):
                sheet = str(cuenta)[-25:] or "Cuenta"
                build_holistor_import(df_cuenta).to_excel(writer, sheet_name=sheet, index=False)
        else:
            build_holistor_import(df_mov).to_excel(writer, sheet_name="Holistor", index=False)

        for ws in writer.book.worksheets:
            _style_sheet(ws)

    return output.getvalue()


def make_operational_summary_pdf(df_mov: pd.DataFrame, df_resumen: pd.DataFrame, df_conciliacion: pd.DataFrame | None = None, logo_path=None) -> bytes:
    output = io.BytesIO()
    doc = SimpleDocTemplate(
        output,
        pagesize=A4,
        rightMargin=1.5 * cm,
        leftMargin=1.5 * cm,
        topMargin=1.3 * cm,
        bottomMargin=1.3 * cm,
    )

    styles = getSampleStyleSheet()
    title = ParagraphStyle("AIETitle", parent=styles["Title"], fontName="Helvetica-Bold", fontSize=16, leading=20, alignment=1)
    normal = ParagraphStyle("AIENormal", parent=styles["Normal"], fontName="Helvetica", fontSize=9, leading=12)
    small = ParagraphStyle("AIESmall", parent=styles["Normal"], fontName="Helvetica", fontSize=8, leading=10, textColor=colors.HexColor("#555555"))

    story = []
    story.append(Paragraph("IA Resumen Bancario – Banco Macro", title))
    story.append(Paragraph("Resumen operativo - Últimos Movimientos", normal))
    story.append(Spacer(1, 0.25 * cm))

    cuentas = df_mov["cuenta"].nunique() if "cuenta" in df_mov.columns else 1
    movimientos = len(df_mov)
    fecha_desde = _fmt_date(df_mov["fecha"].min()) if "fecha" in df_mov.columns else ""
    fecha_hasta = _fmt_date(df_mov["fecha"].max()) if "fecha" in df_mov.columns else ""
    creditos = df_mov.loc[df_mov["importe"] > 0, "importe"].sum()
    debitos = -df_mov.loc[df_mov["importe"] < 0, "importe"].sum()
    neto = df_mov["importe"].sum()

    meta = [
        ["Cuentas detectadas", str(cuentas)],
        ["Movimientos detectados", str(movimientos)],
        ["Período detectado", f"{fecha_desde} al {fecha_hasta}"],
        ["Créditos", f"$ {fmt_money(creditos)}"],
        ["Débitos", f"$ {fmt_money(debitos)}"],
        ["Neto", f"$ {fmt_money(neto)}"],
    ]
    table_meta = Table(meta, colWidths=[6 * cm, 10 * cm])
    table_meta.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#EAF2F8")),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CCCCCC")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    story.append(table_meta)
    story.append(Spacer(1, 0.35 * cm))

    if df_conciliacion is not None and not df_conciliacion.empty:
        story.append(Paragraph("Conciliación bancaria", styles["Heading2"]))
        conc_data = [["Cuenta", "Saldo anterior", "Créditos", "Débitos", "Saldo final", "Diferencia", "Estado"]]
        for _, row in df_conciliacion.iterrows():
            conc_data.append([
                str(row.get("Cuenta", "")),
                f"$ {fmt_money(row.get('Saldo anterior calculado', 0))}",
                f"$ {fmt_money(row.get('Créditos', 0))}",
                f"$ {fmt_money(row.get('Débitos', 0))}",
                f"$ {fmt_money(row.get('Saldo final informado', 0))}",
                f"$ {fmt_money(row.get('Diferencia', 0))}",
                str(row.get("Estado", "")),
            ])
        table_conc = Table(conc_data, colWidths=[3.2 * cm, 2.7 * cm, 2.4 * cm, 2.4 * cm, 2.7 * cm, 2.2 * cm, 2.0 * cm])
        table_conc.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("ALIGN", (1, 1), (5, -1), "RIGHT"),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CCCCCC")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F7F7F7")]),
        ]))
        story.append(table_conc)
        story.append(Spacer(1, 0.35 * cm))

    story.append(Paragraph("Resumen operativo", styles["Heading2"]))
    data = [["Concepto", "Importe"]]
    for _, row in df_resumen.iterrows():
        data.append([str(row["Concepto"]), f"$ {fmt_money(row['Importe'])}"])

    table = Table(data, colWidths=[12 * cm, 4 * cm])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 8.5),
        ("ALIGN", (1, 1), (1, -1), "RIGHT"),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CCCCCC")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, colors.HexColor("#F7F7F7")]),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#D9EAF7")),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
    ]))
    story.append(table)
    story.append(Spacer(1, 0.3 * cm))
    story.append(Paragraph(f"Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')} | Herramienta para uso interno AIE San Justo", small))

    doc.build(story)
    return output.getvalue()
